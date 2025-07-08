# Copyright (c) 2025, Antonio Cañada Momblant and contributors

import frappe
import os
import zipfile
import traceback
import logging
from frappe.utils.file_manager import get_file
from frappe.utils import now_datetime, nowdate
from frappe.model.document import Document
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from erpnext_chile_factura.erpnext_chile_sii_integration.utils.xml_processor import procesar_xml_content


# Configurar logger
logger = logging.getLogger("xml_importer")
if not logger.handlers:
    log_path = frappe.utils.get_site_path("logs", "xml_importer.log")
    os.makedirs(os.path.dirname(log_path), exist_ok=True)

    handler = logging.FileHandler(log_path, mode="a", encoding="utf-8")
    formatter = logging.Formatter("[%(asctime)s] %(levelname)s: %(message)s", "%Y-%m-%d %H:%M:%S")
    handler.setFormatter(formatter)
    logger.addHandler(handler)
    logger.setLevel(logging.DEBUG)


class XMLImporter(Document):
    pass


@frappe.whitelist()
def procesar_xml_zip(docname):
    """
    Encola el procesamiento del archivo ZIP y marca como 'En proceso' inmediatamente.
    """
    frappe.db.set_value("XML Importer", docname, "status", "En proceso")
    frappe.enqueue(
        "erpnext_chile_factura.erpnext_chile_sii_integration.doctype.xml_importer.xml_importer.procesar_xml_zip_direct",
        queue='default',
        timeout=600,
        docname=docname
    )


def procesar_xml_zip_direct(docname):
    """
    Proceso real que se ejecuta en background para procesar el ZIP.
    """
    doc = frappe.get_doc("XML Importer", docname)
    logs = []

    try:
        doc.db_set("status", "En proceso")
        doc.db_set("fecha_carga", now_datetime())
        logger.info(f"== Inicio procesamiento XML Importer: {docname} ==")

        # Obtener archivo ZIP
        file_url = doc.archivo_zip
        if not file_url:
            raise Exception("No se ha subido ningún archivo ZIP.")

        file_name, file_content = get_file(file_url)
        zip_path = frappe.utils.get_site_path("private", "xml_imports", f"{docname}.zip")
        os.makedirs(os.path.dirname(zip_path), exist_ok=True)
        with open(zip_path, "wb") as f:
            f.write(file_content)

        # Extraer ZIP
        extract_dir = frappe.utils.get_site_path("private", "xml_imports", f"{docname}_extracted")
        os.makedirs(extract_dir, exist_ok=True)
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(extract_dir)

        # Procesar cada XML
        for filename in os.listdir(extract_dir):
            if not filename.lower().endswith(".xml"):
                msg = f"{filename}: ignorado (no es XML)"
                logs.append(msg)
                logger.info(f"{docname} > {msg}")
                continue

            try:
                xml_path = os.path.join(extract_dir, filename)
                with open(xml_path, "rb") as xml_file:
                    content = xml_file.read()
                mensaje = procesar_xml_content(content, filename)
                logs.append(mensaje)
                logger.info(f"{docname} > {filename}: OK")
            except Exception as e:
                error_msg = f"{filename}: ❌ Error al procesar - {str(e)}"
                logs.append(error_msg)
                logger.error(f"{docname} > {error_msg}")

        # Finalizar exitosamente
        generar_log_excel(docname, logs)
        doc.db_set("status", "Completado")
        logger.info(f"== Fin procesamiento XML Importer: {docname} ✅ ==")

    except Exception as e:
        error_msg = f"❌ Error crítico: {str(e)}"
        logger.exception(f"== Error crítico en {docname} ==")
        doc.db_set("status", "Error")
        frappe.log_error(error_msg, "Error en procesamiento XML Importer")


def generar_log_excel(docname, logs):
    """
    Genera un archivo Excel con los resultados del procesamiento y lo adjunta al XML Importer.
    """
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Log Importación XML"

    # Encabezados
    ws.append(["Archivo", "Resultado"])

    # Agregar contenido
    for line in logs:
        if ":" in line:
            archivo, resultado = line.split(":", 1)
        else:
            archivo, resultado = "", line
        ws.append([archivo.strip(), resultado.strip()])

    # Ajustar anchos de columna
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        letra = get_column_letter(column_cells[0].column)
        ws.column_dimensions[letra].width = max_len + 4

    # Ruta y nombre
    fecha = nowdate().replace("-", "")
    excel_filename = f"{fecha}_LOG-IMPORTADOR-{docname}.xlsx"
    carpeta_logs = frappe.utils.get_site_path("private", "files", "xml_logs")
    os.makedirs(carpeta_logs, exist_ok=True)
    full_path = os.path.join(carpeta_logs, excel_filename)

    # Guardar archivo
    wb.save(full_path)

    # Adjuntar a ERPNext
    frappe.get_doc({
        "doctype": "File",
        "file_name": excel_filename,
        "file_url": f"/private/files/xml_logs/{excel_filename}",
        "attached_to_doctype": "XML Importer",
        "attached_to_name": docname,
        "is_private": 1
    }).insert(ignore_permissions=True)
